import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from ttkbootstrap import Style
from PIL import Image, ImageTk
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
import os
from datetime import datetime
import shutil
import zipfile
from bcf.bcfxml import load

class BCFConverter(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("BCF para Excel Converter")
        self.geometry("500x600")

        style = Style(theme='flatly')
        self.configure(bg=style.colors.light)

        self.create_main_page()
        self.create_info_page()

        self.main_page.pack(fill=tk.BOTH, expand=True)

    def create_main_page(self):
        self.main_page = ttk.Frame(self)

        header = ttk.Frame(self.main_page)
        header.pack(pady=20)

        ttk.Label(header, text="BCF para Excel", font=("Helvetica", 24, "bold")).pack()
        ttk.Label(header, text="Powered by Luís Fernando", font=("Helvetica", 12)).pack()

        self.bcf_file_var = tk.StringVar()
        self.create_file_input(self.main_page, "Arquivo BCF:", self.bcf_file_var, command=self.browse_bcf)

        self.logo_file_var = tk.StringVar()
        self.create_file_input(self.main_page, "Logo da Empresa:", self.logo_file_var, command=self.browse_logo)

        self.save_location_var = tk.StringVar()
        self.create_file_input(self.main_page, "Pasta para Salvar:", self.save_location_var, command=self.browse_save_location)

        ttk.Button(self.main_page, text="Próximo", style="success.TButton", command=self.show_info_page).pack(pady=20)

    def create_info_page(self):
        self.info_page = ttk.Frame(self)

        header = ttk.Frame(self.info_page)
        header.pack(pady=20)

        ttk.Label(header, text="BCF para Excel", font=("Helvetica", 24, "bold")).pack()
        ttk.Label(header, text="Powered by Luís Fernando", font=("Helvetica", 12)).pack()

        self.project_name_var = tk.StringVar()
        self.create_input_group(self.info_page, "Nome do Projeto:", self.project_name_var, readonly=True)

        self.responsible_var = tk.StringVar()
        self.create_input_group(self.info_page, "Responsável:", self.responsible_var)

        self.client_var = tk.StringVar()
        self.create_input_group(self.info_page, "Cliente:", self.client_var)

        self.stage_var = tk.StringVar()
        self.create_input_group(self.info_page, "Etapa:", self.stage_var)

        self.city_var = tk.StringVar()
        self.create_input_group(self.info_page, "Cidade:", self.city_var)

        self.date_var = tk.StringVar()
        date_frame = ttk.Frame(self.info_page)
        date_frame.pack(fill=tk.X, padx=20, pady=5)
        ttk.Label(date_frame, text="Data:").pack(side=tk.LEFT)
        ttk.Entry(date_frame, textvariable=self.date_var).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 5))
        ttk.Button(date_frame, text="HOJE", command=self.set_today).pack(side=tk.RIGHT)

        ttk.Button(self.info_page, text="Converter", style="success.TButton", command=self.convert).pack(pady=20)
        ttk.Button(self.info_page, text="Voltar", command=self.show_main_page).pack(pady=5)

    def create_input_group(self, parent, label_text, variable, readonly=False):
        frame = ttk.Frame(parent)
        frame.pack(fill=tk.X, padx=20, pady=5)
        ttk.Label(frame, text=label_text).pack(anchor=tk.W)
        ttk.Entry(frame, textvariable=variable, state='readonly' if readonly else 'normal').pack(fill=tk.X)

    def create_file_input(self, parent, label_text, variable, command):
        frame = ttk.Frame(parent)
        frame.pack(fill=tk.X, padx=20, pady=5)
        ttk.Label(frame, text=label_text).pack(anchor=tk.W)
        ttk.Entry(frame, textvariable=variable, state='readonly').pack(side=tk.LEFT, expand=True, fill=tk.X)
        ttk.Button(frame, text="Procurar", command=command).pack(side=tk.RIGHT)

    def browse_bcf(self):
        filename = filedialog.askopenfilename(filetypes=[("BCF files", "*.bcf")])
        if filename:
            self.bcf_file_var.set(filename)
            self.project_name_var.set(self.get_project_name_from_bcf(filename))

    def browse_logo(self):
        filename = filedialog.askopenfilename(filetypes=[("Image files", "*.png;*.jpg;*.jpeg")])
        if filename:
            self.logo_file_var.set(filename)

    def browse_save_location(self):
        directory = filedialog.askdirectory()
        if directory:
            self.save_location_var.set(directory)

    def show_info_page(self):
        self.main_page.pack_forget()
        self.info_page.pack(fill=tk.BOTH, expand=True)

    def show_main_page(self):
        self.info_page.pack_forget()
        self.main_page.pack(fill=tk.BOTH, expand=True)

    def set_today(self):
        self.date_var.set(datetime.now().strftime("%d/%m/%Y"))

    def convert(self):
        bcf_file = self.bcf_file_var.get()
        xlsx_file = os.path.join(self.save_location_var.get(), f"{self.project_name_var.get()}.xlsx")
        logo_path = self.logo_file_var.get()
        project_name = self.project_name_var.get()
        responsavel = self.responsible_var.get()
        cliente = self.client_var.get()
        etapa = self.stage_var.get()
        cidade = self.city_var.get()
        data = self.date_var.get()

        if not all([bcf_file, xlsx_file, logo_path, project_name, responsavel, cliente, etapa, cidade, data]):
            messagebox.showerror("Erro", "Por favor, preencha todos os campos.")
            return

        # Chama a função de processamento BCF para XLSX
        self.process_bcf(bcf_file, xlsx_file, logo_path, project_name, responsavel, cliente, etapa, cidade, data)

    def get_project_name_from_bcf(self, bcf_file):
        try:
            with load(bcf_file) as bcfxml:
                project = bcfxml.project
                return project.name if project.name else "Projeto"
        except:
            return "Sem Título"

    def process_bcf(self, bcf_file, xlsx_file, logo_path, project_name, responsavel, cliente, etapa, cidade, data):
        extract_folder = os.path.join(os.path.dirname(bcf_file), "extracted_bcf")
        
        # Extrai o conteúdo do BCF
        with zipfile.ZipFile(bcf_file, 'r') as zip_ref:
            zip_ref.extractall(extract_folder)

        with load(bcf_file) as bcfxml:
            topics = bcfxml.topics

            topic_titles, labels, creation_dates, assigned_tos, stages, descriptions, comments, snapshots = [], [], [], [], [], [], [], []

            # Processa cada tópico no arquivo BCF
            for topic_guid, topic_handler in topics.items():
                try:
                    topic = topic_handler.topic
                    topic_titles.append(topic.title)
                    labels.append(topic.labels)
                    creation_dates.append(topic.creation_date)
                    assigned_tos.append(topic.assigned_to)
                    stages.append(topic.stage)
                    descriptions.append(topic.description)

                    comments_list = []
                    for comment in topic_handler.comments:
                        comments_list.append({
                            'comment': comment.comment,
                            'author': comment.author
                        })
                    comments.append(comments_list)

                    # Processa snapshots (imagens)
                    topic_folder = os.path.join(extract_folder, topic_guid)
                    snapshot_path = next((os.path.join(topic_folder, f) for f in os.listdir(topic_folder) if f.lower().endswith(('.png', '.jpg', '.jpeg'))), None)
                    snapshots.append(snapshot_path)

                except Exception as e:
                    print(f"Erro ao processar tópico {topic_guid}: {e}")

        wb = Workbook()
        ws = wb.active

        # Adiciona o cabeçalho personalizado
        self.add_custom_header(ws, logo_path, project_name, responsavel, cliente, etapa, cidade, data)

        # Adiciona os dados dos tópicos ao Excel
        for i, (topic_title, label, creation_date, assigned_to, stage, description, comments_list, snapshot_path) in enumerate(zip(topic_titles, labels, creation_dates, assigned_tos, stages, descriptions, comments, snapshots), start=10):
            comments_str = self.format_comments(comments_list)
            creation_date_str = datetime.strftime(creation_date.to_datetime(), "%Y-%m-%d %H:%M:%S")

            ws.cell(row=i, column=1, value=topic_title)
            ws.cell(row=i, column=2, value=', '.join(label))
            ws.cell(row=i, column=3, value=creation_date_str)
            ws.cell(row=i, column=4, value=assigned_to)
            ws.cell(row=i, column=5, value=stage)
            ws.cell(row=i, column=6, value=description)
            ws.cell(row=i, column=8, value=comments_str)

            for col in range(1, 9):
                ws.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical='top')

            # Ajusta a altura da linha para acomodar a imagem e o texto
            ws.row_dimensions[i].height = 128  # Altura em pontos

            if snapshot_path and os.path.exists(snapshot_path):
                img = ExcelImage(snapshot_path)
                # Ajusta o tamanho da imagem para caber na célula
                img.width = 256  # Largura em pixels
                img.height = 128  # Altura em pixels
                # Posiciona a imagem na célula G
                img.anchor = f'G{i}'
                ws.add_image(img)
            else:
                ws.cell(row=i, column=7, value="Nenhum snapshot disponível")

        # Salva o arquivo Excel
        wb.save(xlsx_file)
        
        # Limpa arquivos temporários
        shutil.rmtree(extract_folder)

        messagebox.showinfo("Sucesso", f"Conversão concluída com sucesso! Arquivo salvo em: {xlsx_file}")

    def add_custom_header(self, ws, logo_path, project_name, responsavel, cliente, etapa, cidade, data):
        # Adicione o logo
        if os.path.exists(logo_path):
            logo = ExcelImage(logo_path)
            logo.width = 200
            logo.height = 50
            ws.add_image(logo, 'A1')

        # Adicione o nome do projeto
        ws.merge_cells('F1:H2')
        project_cell = ws['F1']
        project_cell.value = project_name
        project_cell.font = Font(bold=True, size=14)
        project_cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

        # Adicione as informações do projeto
        info = [
            ('RESPONSÁVEL', responsavel),
            ('CLIENTE', cliente),
            ('ETAPA', etapa),
            ('CIDADE', cidade),
            ('DATA', data)
        ]
        for i, (label, value) in enumerate(info, start=3):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'].font = Font(bold=True)

        # Adiciona linha horizontal
        for col in range(1, 9):
            ws.cell(row=8, column=col).border = Border(bottom=Side(style='thin'))

        # Adicione os cabeçalhos da tabela
        headers = ['TÍTULO', 'DISCIPLINA', 'DATA', 'AUTOR', 'FASE', 'DESCRIÇÃO', 'IMAGEM', 'COMENTÁRIOS']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Ajusta a largura das colunas
        column_widths = [20, 15, 18, 20, 15, 50, 40, 60]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def format_comments(self, comments_list):
        formatted_comments = []
        for comment_data in comments_list:
            comment_text = f"Autor: {comment_data['author']}\nComentário: {comment_data['comment']}"
            formatted_comments.append(comment_text)
        return "\n\n".join(formatted_comments)

if __name__ == "__main__":
    app = BCFConverter()
    app.mainloop()
